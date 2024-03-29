from Modulos.UnidadProductiva.models import UnidadProductiva
from Modulos.UnidadNegocio.models import UnidadNegocio
from Modulos.CentroCostos.models import SubCentroCosto,CentroCosto
from Modulos.Usuario.models import Usuario
from django.db.models import Sum
from django.db.models import Q
from django.db import models
from django.conf import settings
from openpyxl import load_workbook
import os
import locale
import uuid
import openpyxl
from django.utils import timezone
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from django.http import HttpResponse
import pyexcel as pe
from fpdf import FPDF
from django.http import FileResponse
from fpdf import FPDF
import boto3



# Create your models here.

class TipoMovimiento(models.TextChoices):
    INGRESO = 'IN', 'Ingreso-s'
    EGRESO = 'OUT', 'Egreso-s'

class Movimiento(models.Model):
    unidad_productiva = models.ForeignKey(UnidadProductiva, on_delete=models.CASCADE,null=True,blank=True)
    numero_documento = models.CharField(max_length=50,null=True,blank=True)
    nombre_proveedor = models.CharField(max_length=50,null=True,blank=True)
    sub_centro_costo = models.ForeignKey(SubCentroCosto, on_delete=models.CASCADE,null=True,blank=True)
    numero_factura = models.CharField(max_length=50,null=True,blank=True)
    tipo_documento = models.CharField(max_length=50,null=True,blank=True)
    factura = models.BooleanField(default=False,null=True,blank=True)
    estado = models.CharField(max_length=10,null=True,blank=True)
    accion = models.CharField(max_length=200,null=True,blank=True)
    uuid = models.SlugField(blank=True)
    fecha_registro = models.DateTimeField(default=timezone.now, null=True, blank=True, editable=True)
    fecha_factura = models.DateField(null=True, blank=True, editable=True)
    fecha_modificacion = models.DateTimeField(auto_now_add=True,null=True,blank=True)
    #concepto = models.CharField(max_length=200,null=True,blank=True)
    concepto = models.TextField(null=True,blank=True)
    valor = models.FloatField(default=0)
    comprobante_factura = models.FileField(upload_to='comprobantes/',null=True,blank=True)
    ingreso_bancario = models.BooleanField(default=False,null=True,blank=True)
    negociacion = models.CharField(max_length=255,null=True,blank=True)
    usuario_presupuesto = models.ForeignKey(Usuario, on_delete=models.CASCADE,null=True,blank=True)
    usuario_admin_ingreso = models.ForeignKey(Usuario, on_delete=models.CASCADE,null=True,blank=True,related_name='usuario_admin_ingreso')

    tipo_ingreso = models.CharField(
        default=TipoMovimiento.INGRESO,
        max_length=50,
    )
    



    def __str__(self):
        fecha_registro = self.fecha_registro.astimezone(timezone.get_current_timezone())
        fecha_formateada_registro = fecha_registro.strftime('%d/%m/%Y %H:%M')
        return f"{fecha_formateada_registro} -{self.tipo_ingreso} - {self.concepto} - {self.valor}"
    
    def gen_uid(self):
        self.uid = uuid.uuid4()
        self.save()
    
    import os
from django.conf import settings
import boto3

def subir_archivo_s3(self, nombre_archivo, archivo):
    # Guardar el archivo temporalmente en el sistema de archivos
    ruta_temporal = os.path.join(settings.MEDIA_ROOT, 'temp', nombre_archivo)
    with open(ruta_temporal, 'wb+') as destination:
        for chunk in archivo.chunks():
            destination.write(chunk)

    # Cargar el archivo desde el sistema de archivos a AWS S3
    s3 = boto3.client('s3', aws_access_key_id=settings.AWS_ACCESS_KEY_ID, aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY)
    s3.upload_file(ruta_temporal, settings.AWS_STORAGE_BUCKET_NAME, nombre_archivo)

    # Eliminar el archivo temporal
    os.remove(ruta_temporal)


def export_to_excel(queryset,to_pdf=False):
    # Obtener los objetos de Django
    #queryset = Movimiento.objects.all()

    # Crear el libro de trabajo de Excel y la hoja de cálculo
    workbook = Workbook()
    workbook.iso_dates = True
    worksheet = workbook.active

    # Escribir los encabezados de las columnas
    headers = ['Fecha de Registro', 'Fecha Modificación', 'Tipo de Ingreso', 'Unidad Productiva','Usuario Ingreso Caja', 'Concepto', 'Valor','Centro Costo', 'SubCentro Costo', 'Factura', 'Numero de Factura', 'Fecha de Factura','Nombre del Proveedor','Tipo de Documento Proveedor','Nro Documento Proveedor', 'Estado', 'Accion', 'Comprobante Factura']
    for col_num, header in enumerate(headers, 1):
        worksheet.cell(row=1, column=col_num, value=header)

    # Escribir los datos del objeto
    for row_num, obj in enumerate(queryset, 2):
        if not obj.unidad_productiva:
            unidad_productiva_nombre = 'N/A'
        else:
            unidad_productiva_nombre =  obj.unidad_productiva.nombre
        if obj.tipo_ingreso=='IN':
            if obj.ingreso_bancario:
                tipo_ingreso = 'Ingreso Bancario'
            else:
                tipo_ingreso = 'Ingreso de caja'
            if obj.accion == 'Reducción de Caja':
                tipo_ingreso = 'Reducción de Caja'
        elif obj.tipo_ingreso=='OUT':
            if obj.ingreso_bancario:
                tipo_ingreso = 'Egreso Bancario'
            else:
                tipo_ingreso = 'Egreso de caja'
        else:
            tipo_ingreso = 'N/A'

        if not obj.sub_centro_costo:
            centro_costo = 'N/A'
            sub_centro_costo_nombre = 'N/A'
        else:
            sub_centro_costo_nombre = obj.sub_centro_costo.nombre
            centro_costo = CentroCosto.objects.filter(subcentro=obj.sub_centro_costo).first().nombre
        
        if not obj.usuario_presupuesto:
            usuario_presupuesto = 'N/A'
        else:
            usuario_presupuesto = obj.usuario_presupuesto.nombre + ' ' + obj.usuario_presupuesto.apellido
        
        if not obj.nombre_proveedor:
            proveedor_nombre = 'N/A'
            tipo_documento = 'N/A'
            nro_documento = 'N/A'
        else:
            proveedor_nombre = obj.nombre_proveedor
            tipo_documento = obj.tipo_documento
            nro_documento = obj.numero_documento
            
        fecha_registro = obj.fecha_registro.astimezone(timezone.get_current_timezone())
        fecha_formateada_registro = fecha_registro.strftime('%d/%m/%Y %H:%M')    
        worksheet.cell(row=row_num, column=1, value=fecha_formateada_registro)
        fecha_modificacion = obj.fecha_modificacion.astimezone(timezone.get_current_timezone())
        fecha_formateada_registro = fecha_modificacion.strftime('%d/%m/%Y %H:%M')    
        worksheet.cell(row=row_num, column=2, value=fecha_formateada_registro)
        worksheet.cell(row=row_num, column=3, value=tipo_ingreso)
        worksheet.cell(row=row_num, column=4, value=unidad_productiva_nombre)
        worksheet.cell(row=row_num, column=5, value=usuario_presupuesto)
        worksheet.cell(row=row_num, column=6, value=obj.concepto)
        worksheet.cell(row=row_num, column=7, value=obj.valor)
        worksheet.cell(row=row_num, column=8, value=centro_costo)
        worksheet.cell(row=row_num, column=9, value=sub_centro_costo_nombre)
        worksheet.cell(row=row_num, column=10, value=obj.factura)
        worksheet.cell(row=row_num, column=11, value=obj.numero_factura)
        worksheet.cell(row=row_num, column=12, value=obj.fecha_factura)
        worksheet.cell(row=row_num, column=13, value=proveedor_nombre)
        worksheet.cell(row=row_num, column=14, value=tipo_documento)
        worksheet.cell(row=row_num, column=15, value=nro_documento)
        
        worksheet.cell(row=row_num, column=16, value=obj.estado)
        worksheet.cell(row=row_num, column=17, value=obj.accion)
    
    if to_pdf:
        workbook.save('archivo_excel.xlsx')
        return convert_xlsx_to_pdf('archivo_excel.xlsx', 'pdf_file.pdf')
    else:
        # Crear la respuesta del archivo Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=archivo.xlsx'
        workbook.save(response)
        return response




def convert_xlsx_to_pdf(xlsx_file, pdf_file):
    # Load the XLSX file
    workbook = load_workbook(xlsx_file)

    # Select the active worksheet
    worksheet = workbook.active

    # Create the PDF object
    pdf = FPDF()

    # Add a page to the PDF
    pdf.add_page()

    # Iterate over the rows of the XLSX and add them to the PDF
    for row in worksheet.iter_rows(values_only=True):
        line = '\t'.join(str(cell) for cell in row)
        pdf.cell(0, 10, txt=line, ln=True)

    # Guardar el archivo PDF
    pdf.output(pdf_file)
     # Crear una instancia de FileResponse con el archivo PDF
    response = FileResponse(open(pdf_file, 'rb'), content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="archivo.pdf"'

    return response


    

def get_estado_caja(user,usuario_admin=None):
    union_query = Q()

    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')
    filtros_in = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&Q(usuario_presupuesto=user)&Q(ingreso_bancario=False)&union_query
    filtros_out = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioRegistro=user)&Q(ingreso_bancario=False)&Q(usuario_presupuesto=user)&union_query
    filtros_in_ba = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioRegistro=user)&Q(ingreso_bancario=True)&union_query
    filtros_out_ba = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioRegistro=user)&Q(ingreso_bancario=True)&union_query
    if user.groups.filter(name__in=['Observador']).exists():
        filtros_in = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&(Q(unidad_productiva__usuarioConsulta=user)| Q(usuario_presupuesto=user))&Q(ingreso_bancario=False)&union_query
        filtros_out = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioConsulta=user)&Q(ingreso_bancario=False)&union_query
        filtros_in_ba = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioConsulta=user)&Q(ingreso_bancario=True)&union_query
        filtros_out_ba = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioConsulta=user)&Q(ingreso_bancario=True)&union_query
    if user.groups.filter(name__in=['Bancario']).exists():
        filtros_in = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&(Q(unidad_productiva__usuarioBancario=user)| Q(usuario_presupuesto=user))&Q(ingreso_bancario=False)&union_query
        filtros_out = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioBancario=user)&Q(ingreso_bancario=False)&union_query
        filtros_in_ba = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioBancario=user)&Q(ingreso_bancario=True)&union_query
        filtros_out_ba = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&Q(unidad_productiva__usuarioBancario=user)&Q(ingreso_bancario=True)&union_query
    
    if usuario_admin:
        unidades_negocio = UnidadNegocio.objects.filter(admin=usuario_admin).all()
        for unidad_negocio in unidades_negocio:
            condicion = Q(unidad_productiva__in=unidad_negocio.unidades_productivas.all())
            union_query |= condicion
        union_query |= Q(usuario_admin_ingreso=usuario_admin)
        filtros_in = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&Q(ingreso_bancario=False)&Q(usuario_presupuesto=user)

    ingresos =  Movimiento.objects.distinct().filter(filtros_in).aggregate(Sum('valor'))['valor__sum']
    egresos = Movimiento.objects.distinct().filter(filtros_out).aggregate(Sum('valor'))['valor__sum']
    ingresos_ba = Movimiento.objects.distinct().filter(filtros_in_ba).aggregate(Sum('valor'))['valor__sum']
    egresos_ba = Movimiento.objects.distinct().filter(filtros_out_ba).aggregate(Sum('valor'))['valor__sum']

    if ingresos == None:
        ingresos = 0
    if egresos == None:
        egresos = 0

    user.presupuesto=ingresos
    user.save()

    ingresos_f = locale.currency(ingresos, symbol=True, grouping=True)
    egresos_f = locale.currency(egresos, symbol=True, grouping=True)
    diferencia = ingresos - egresos
    diferencia_f = locale.currency(diferencia, symbol=True, grouping=True)

    if ingresos_ba == None:
        ingresos_ba = 0
    if egresos_ba == None:
        egresos_ba = 0
    ingresos_ba_f = locale.currency(ingresos_ba, symbol=True, grouping=True)
    egresos_ba_f = locale.currency(egresos_ba, symbol=True, grouping=True)
    diferencia_ba = ingresos_ba - egresos_ba
    diferencia_ba_f = locale.currency(diferencia_ba, symbol=True, grouping=True)

    return diferencia_f, ingresos_f, egresos_f, diferencia_ba_f, ingresos_ba_f, egresos_ba_f

def get_estado_caja_admin(user,unidad_productiva=None):
    locale.setlocale(locale.LC_ALL, 'en_US.UTF-8')

    unidades_negocio = UnidadNegocio.objects.filter(admin=user).all()
    # Inicializa un objeto Q vacío
    union_query = Q()
    for unidad_negocio in unidades_negocio:
        condicion = Q(unidad_productiva__in=unidad_negocio.unidades_productivas.all()) 
        union_query |= condicion
    
    union_query |= Q(unidad_productiva__usuarioRegistro=user)
    union_query |= Q(usuario_presupuesto=user)
    union_query |= Q(usuario_admin_ingreso=user)


    if user.groups.filter(name__in=['Auditor']).exists():
        unidad_productivas = UnidadProductiva.objects.filter(usuarioAuditor=user).all()
        usuarios_registro = unidad_productivas.values('usuarioRegistro')    
        union_query |= Q(usuario_presupuesto__in=usuarios_registro)
        union_query |= Q(unidad_productiva__usuarioAuditor=user)
    
    filtros_in = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&Q(ingreso_bancario=False)
    filtros_out = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&union_query&Q(ingreso_bancario=False)
    filtros_in_ba = Q(tipo_ingreso='IN')&Q(estado='Aprobado')&union_query&Q(ingreso_bancario=True)
    filtros_out_ba = Q(tipo_ingreso='OUT')&Q(estado='Aprobado')&union_query&Q(ingreso_bancario=True)

    ingresos = Movimiento.objects.filter(filtros_in).distinct().aggregate(Sum('valor'))['valor__sum']
    egresos = Movimiento.objects.filter(filtros_out).distinct().aggregate(Sum('valor'))['valor__sum']
    ingresos_ba = Movimiento.objects.filter(filtros_in_ba).distinct().aggregate(Sum('valor'))['valor__sum']
    egresos_ba = Movimiento.objects.filter(filtros_out_ba).distinct().aggregate(Sum('valor'))['valor__sum']

    if ingresos == None:
        ingresos = 0
    if egresos == None:
        egresos = 0

    ingresos_f = locale.currency(ingresos, symbol=True, grouping=True)
    egresos_f = locale.currency(egresos, symbol=True, grouping=True)
    diferencia = ingresos - egresos
    diferencia_f = locale.currency(diferencia, symbol=True, grouping=True)

    if ingresos_ba == None:
        ingresos_ba = 0
    if egresos_ba == None:
        egresos_ba = 0
    ingresos_ba_f = locale.currency(ingresos_ba, symbol=True, grouping=True)
    egresos_ba_f = locale.currency(egresos_ba, symbol=True, grouping=True)
    diferencia_ba = ingresos_ba - egresos_ba
    diferencia_ba_f = locale.currency(diferencia_ba, symbol=True, grouping=True)

    return diferencia_f, ingresos_f, egresos_f, diferencia_ba_f, ingresos_ba_f, egresos_ba_f



def get_movimientos_usuario(usuario):
    if usuario.groups.filter(name__in=['Administrador']).exists():
        querysets = []
        unidades_negocio = UnidadNegocio.objects.filter(admin=usuario).all()
        # Inicializa un objeto Q vacío
        union_query = Q()
        for unidad_negocio in unidades_negocio:
            condicion = Q(unidad_productiva__in=unidad_negocio.unidades_productivas.all()) 
            union_query |= condicion

        
        union_query |= Q(usuario_admin_ingreso=usuario)
        return  Movimiento.objects.all()
    elif usuario.groups.filter(name__in=['Auditor']).exists():
        unidades_negocio = UnidadNegocio.objects.all()
        # Inicializa un objeto Q vacío
        union_query = Q()
        for unidad_negocio in unidades_negocio:
            condicion = Q(unidad_productiva__in=unidad_negocio.unidades_productivas.all()) 
            union_query |= condicion
        #usuarios_registro = Usuario.objects.filter(unidadproductiva__usuarioAuditor=usuario).values('unidadproductiva__usuarioRegistro') 
        unidad_productivas = UnidadProductiva.objects.filter(usuarioAuditor=usuario).all()
        usuarios_registro = unidad_productivas.values('usuarioRegistro')     
        union_query |= Q(usuario_presupuesto__in=usuarios_registro)
        union_query |= Q(usuario_presupuesto=usuario)
        union_query |= Q(unidad_productiva__usuarioAuditor=usuario)
        union_query |= Q(unidad_productiva__usuarioAuditor=usuario)
        union_query |= Q(usuario_admin_ingreso=usuario)
        return  Movimiento.objects.filter(union_query).all()

    else:
        if usuario.groups.filter(name__in=['Comun','Bancario','Observador']).exists():
            filters = Q (unidad_productiva__usuarioRegistro=usuario) | Q(usuario_presupuesto=usuario)|Q(unidad_productiva__usuarioBancario=usuario)|Q(unidad_productiva__usuarioConsulta=usuario) &Q(usuario_presupuesto=usuario)
            movimientos = Movimiento.objects.filter(filters).all()
        elif usuario.groups.filter(name__in=['Observador']).exists():
            filters = Q (unidad_productiva__usuarioConsulta=usuario)
            movimientos = Movimiento.objects.filter(filters).all()
        else:
            movimientos = Movimiento.objects.all()
        return movimientos.filter(Q(usuario_presupuesto=usuario)).all()

def get_movimientos(usuario,unidad_productiva=None):
    if usuario.groups.filter(name__in=['Administrador','Auditor']).exists():
        disponible,ingreso,egreso, disponible_ba,ingreso_ba,egreso_ba = get_estado_caja_admin(usuario,unidad_productiva)
        return  disponible,ingreso,egreso, disponible_ba,ingreso_ba,egreso_ba
    elif usuario.groups.filter(name='Comun').exists():
        unidad_productiva = UnidadProductiva.objects.filter(usuarioRegistro=usuario).first()
        disponible,ingreso,egreso, disponible_ba,ingreso_ba,egreso_ba = get_estado_caja(usuario)
        return disponible,ingreso,egreso,disponible_ba,ingreso_ba,egreso_ba
    else:
        disponible,ingreso,egreso, disponible_ba,ingreso_ba,egreso_ba = get_estado_caja_admin(usuario)
        return disponible,ingreso,egreso, disponible_ba,ingreso_ba,egreso_ba
   












class Comentario(models.Model):
    movimiento = models.ForeignKey(Movimiento, on_delete=models.CASCADE,null=True,blank=True)
    usuario = models.ForeignKey(Usuario, on_delete=models.CASCADE,null=True,blank=True)
    comentario = models.CharField(max_length=200,null=True,blank=True)
    fecha_registro = models.DateTimeField(auto_created=True,auto_now_add=True,null=True,blank=True)
    def __str__(self):
        return f"{self.usuario} - {self.comentario}"
    

